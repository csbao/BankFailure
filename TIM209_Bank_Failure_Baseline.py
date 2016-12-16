#!/usr/bin/env python

import os
import openpyxl

def read_file(filename):
    vals1 = openpyxl.load_workbook(filename)
    sheetnames = vals1.get_sheet_names()
    i = 0
    # print sheetnames
    print (len(sheetnames))
    sheets = [None] * len(sheetnames)
    for name in sheetnames:
        # print (str(i + 1), ". ", name)
        sheets[i] = vals1.get_sheet_by_name(name)
        # print sheets[i]
        i += 1

    print (len(sheets))
    return sheets


def process_sheets(sheets, column, start_row, end_row, filename, label):
    print (len(sheets))
    i = 1
    for sheet in sheets:
        # print sheet
        outfile = open(filename, 'a+')
        print (filename, "File Opened.")
        if os.stat(filename).st_size == 0:
            # with open(filename, 'w') as outfile:
            #     print "File Opened."
            for rows in sheet[str(str(column) + str(start_row)):str(str(column) + str(end_row))]:
                for cell in rows:
                    if cell.value == "#N/A N/A":
                        outfile.write(str(label) + " " + str(i) + ":" + str(0))
                        outfile.write("\n")
                    else:
                        outfile.write(str(label) + " " + str(i) + ":" + str(cell.value))
                        outfile.write("\n")
            # outfile.close()
        else:
            cells = []
            for rows in sheet[str(str(column) + str(start_row)):str(str(column) + str(end_row))]:
                for cell in rows:
                    if cell.value == "#N/A N/A":
                        cells.append(0)
                    else:
                        cells.append(cell.value)

            outfile.seek(0)
            lines = outfile.readlines()
            outfile.close()
            j = 0
            outfile = open(filename, 'w')
            for line in lines:
                if j <= (end_row - start_row):
                    outfile.write('%s%s%s%s%s\n' % (line.rstrip('\n'), " ", str(i), ":", str(cells[j])))
                    # outfile.write("\n")
                    j += 1
                else:
                    break
        outfile.close()
        print ("File closed")
        i += 1


def process_sheets_failed_banks(sheets, columns, start_row, end_row, filename, label):
    print (len(sheets))
    i = 1
    k = 0
    for sheet in sheets:
        # print sheet
        outfile = open(filename, 'a+')
        print ("File opened")
        count = start_row
        if os.stat(filename).st_size == 0:
            for column in columns:
                cell_val = sheet[str(column) + str(count)].value
                if cell_val == "#N/A N/A":
                    outfile.write(str(label) + " " + str(i) + ":" + "0")
                else:
                    outfile.write(str(label) + " " + str(i) + ":" + str(cell_val))
                outfile.write("\n")
                count += 1
            # outfile.close()
        else:
            cells = []
            for column in columns:
                cell_val = sheet[str(column) + str(count)].value
                if cell_val == "#N/A N/A":
                    cells.append(0)
                else:
                    cells.append(cell_val)
                count += 1
                print (cell_val)

            outfile.seek(0)
            lines = outfile.readlines()
            outfile.close()
            j = 0
            outfile = open(filename, 'w')
            for line in lines:
                if j <= (end_row - start_row):
                    outfile.write('%s%s%s%s%s\n' % (line.rstrip('\n'), " ", str(i), ":", str(cells[j])))
                    # outfile.write("\n")
                    j += 1
                else:
                    break
        outfile.close()
        print ("File closed")
        i += 1


def split_into_training_test(input_file, training_file, test_file):
    infile = open(input_file, 'r')
    train_f = open(training_file, 'a')
    test_f = open(test_file, 'a')
    i = 0
    for line in infile:
        if (i % 5 == 0):
            test_f.write(line)
            # test_f.write('\n')
        else:
            train_f.write(line)
            # train_f.write('\n')
        i += 1
    infile.close()
    train_f.close()
    test_f.close()


def file_length(filename):
    i = 0
    with open(filename) as f:
        for i, l in enumerate(f):
            pass

    return i + 1


if __name__ == '__main__':

    data_file = "BankList_Values_20161124.xlsx"
    data_file2 = "BankList_Values_20161115.xlsx"

    sheets = read_file(data_file)
    sheets2 = read_file(data_file2)
    sheets_total = sheets + sheets2
    print (len(sheets_total))

    columns = ['AK', 'AI', 'BK', 'AM', 'AO', 'BG', 'AJ', 'BC', 'AJ', 'AR', 'AF', 'AY', 'AK', 'AJ', 'AB', 'BE', 'AN',
              'AJ', 'AJ', 'AC', 'AI', 'AN', 'AV', 'AJ', 'AS', 'AL', 'AM', 'AP', 'AD', 'AP', 'AR', 'AQ', 'AI', 'AM']

    label_failed = 0
    label_active = 1
    failed_file = "failed_banks.txt"
    process_sheets_failed_banks(sheets_total, columns, 3, 36, failed_file, label_failed)

    active_file = "active_banks.txt"
    column_used_active_banks = 'Z'
    process_sheets(sheets_total, column_used_active_banks, 38, 167, active_file, label_active)
    ##########################################
    # I combined the "failed_file" and "active_file" manually,
    # and then used it as "data_txtfile" to run the "split_into_training_test" function.
    ##########################################
    data_txtfile = "banks_Z_1colBeforeFailed.txt"
    training_file = "training_data_Z_1col.txt"
    test_file = "test_data_Z_1col.txt"
    split_into_training_test(data_txtfile, training_file, test_file)

